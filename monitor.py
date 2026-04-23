#!/usr/bin/env python3
"""
Monitoring : Uptime + Response time + SSL + NDD + Core Web Vitals + Stack + Backup
- Lit l'inventaire depuis inventaire_monitoring.xlsx (read_only)
- Ecrit les resultats dans results.json + l Excel (colonnes K-P)
- Stocke l'historique dans history.db (SQLite) pour les tendances
- Envoie un email a support@selnhelp.fr en cas d'alerte
- Lance PageSpeed Insights uniquement si --psi (cron hebdo) ou PSI_API_KEY defini
"""

import json
import logging
import os
import smtplib
import sys
import warnings
from datetime import datetime
from email.mime.text import MIMEText
from pathlib import Path

from openpyxl import load_workbook

from checks import (
    check_uptime,
    check_ssl,
    check_domain,
    check_pagespeed,
    check_stack,
    load_backup_status,
)
import history

warnings.filterwarnings("ignore")

# ── Chemins ───────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "inventaire_monitoring.xlsx"
RESULTS_JSON = BASE_DIR / "results.json"
HISTORY_DB = BASE_DIR / "history.db"
LOG_PATH = BASE_DIR / "monitor.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler()],
)
log = logging.getLogger(__name__)

# Colonnes Excel (1-indexed)
COL_CLIENT = 1
COL_DOMAINE = 2
COL_URL = 3
COL_HEB_DOM = 4
COL_HEB_SITE = 5
COL_RESP = 7
COL_DATE_EXP = 9


# ── Config depuis l onglet Config ─────────────────────────────────────────────
def load_config() -> dict:
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb["Config"]
        cfg = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and row[1]:
                cfg[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        return cfg
    except Exception as e:
        log.warning(f"Impossible de lire la config Excel ({e}) — valeurs par defaut")
        return {}


# ── Envoi email via SMTP ──────────────────────────────────────────────────────
def _send_email(to_addr, subject: str, body: str, cfg: dict) -> bool:
    smtp_host = os.environ.get("SMTP_HOST", cfg.get("SMTP host", "smtp.office365.com"))
    smtp_port = int(os.environ.get("SMTP_PORT", cfg.get("SMTP port", "587")))
    smtp_user = os.environ.get("SMTP_USER", cfg.get("SMTP user", ""))
    smtp_pass = os.environ.get("SMTP_PASSWORD", cfg.get("SMTP password", ""))
    from_addr = os.environ.get("EMAIL_FROM", smtp_user)

    if not smtp_user or not smtp_pass:
        log.error("SMTP non configure — definissez SMTP_USER et SMTP_PASSWORD")
        return False

    if isinstance(to_addr, str):
        to_list = [a.strip() for a in to_addr.split(",") if a.strip()]
    else:
        to_list = [a.strip() for a in to_addr if a.strip()]

    try:
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = from_addr
        msg["To"] = ", ".join(to_list)
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(from_addr, to_list, msg.as_string())
        log.info(f"Email envoye a {', '.join(to_list)} : {subject}")
        return True
    except Exception as e:
        log.error(f"Echec envoi email : {e}")
        return False


def send_support_email(cfg: dict, subject: str, body: str) -> bool:
    support = os.environ.get("EMAIL_SUPPORT", cfg.get("Email support", "supportcsm@albys.com"))
    axel = os.environ.get("EMAIL_AXEL", cfg.get("Email axel", "axel.dos-santos@albys.com"))
    to_list = list(dict.fromkeys(filter(None, [support, axel])))
    log.info(f"Envoi email vers {', '.join(to_list)}...")
    return _send_email(to_list, subject, body, cfg)


def send_test_report(results: list, cfg: dict, now: str):
    to_addr = os.environ.get("EMAIL_TEST", cfg.get("Email test", "axel.dos-santos@albys.com"))
    subject = f"[MONITORING] Rapport complet — {now}"

    alertes = [r for r in results if not r["up"] or r["ssl_st"] in ("warning", "critical")]
    sains = [r for r in results if r not in alertes]

    lignes = [f"Rapport de monitoring — {now}", "=" * 60, ""]
    lignes.append(f"Total sites : {len(results)}   Alertes : {len(alertes)}")
    lignes.append("")

    if alertes:
        lignes.append("SITES NECESSITANT UNE ATTENTION")
        lignes.append("-" * 40)
        for r in alertes:
            uptime = "UP  " if r["up"] else "DOWN"
            lignes.append(f"[{uptime}] {r['client']} ({r['url']})")
            if not r["up"]:
                lignes.append(f"         Uptime : {r['up_msg']}")
            lignes.append(f"         SSL    : {r['ssl_msg']}")
            lignes.append(f"         NDD    : {r['ndd_msg']}")
            if r.get("response_ms"):
                lignes.append(f"         Temps  : {r['response_ms']} ms")
            psi = r.get("psi") or {}
            if psi.get("score") is not None:
                lignes.append(f"         PSI    : {psi['score']}/100 — LCP {psi.get('lcp_ms')}ms")
            lignes.append("")

    if sains:
        lignes.append("SITES OK")
        lignes.append("-" * 40)
        for r in sains:
            rt = f" [{r['response_ms']}ms]" if r.get("response_ms") else ""
            lignes.append(f"[UP  ] {r['client']} ({r['url']}){rt} — SSL : {r['ssl_msg']}")
        lignes.append("")

    body = "\n".join(lignes)
    log.info(f"Envoi rapport vers {to_addr}...")
    _send_email(to_addr, subject, body, cfg)


# ── Main ──────────────────────────────────────────────────────────────────────
def run(test_mode: bool = False, run_psi: bool = False):
    log.info("=== Demarrage du monitoring ===")
    cfg = load_config()
    timeout = int(cfg.get("Timeout HTTP (secondes)", "10"))
    ssl_warn = int(cfg.get("Alerte SSL (jours avant)", "30"))
    ndd_warn = int(cfg.get("Alerte NDD (jours avant)", "30"))
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Backup status (un seul fetch pour tous les sites)
    backup_index = load_backup_status()
    if backup_index:
        log.info(f"Backup status charge : {len(backup_index)} domaines")
    else:
        log.info("Backup status non disponible (definir BACKUP_STATUS_URL ou BACKUP_STATUS_PATH)")

    # Lecture Excel
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb["Inventaire"]
    except PermissionError:
        log.error("Excel verrouille (ouvert dans Excel ou sync OneDrive en cours). Ferme-le et relance.")
        return
    except Exception as e:
        log.error(f"Impossible de lire l Excel: {e}")
        return

    results = []
    excel_writes = []

    for excel_row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        client = row[COL_CLIENT - 1]
        domaine = row[COL_DOMAINE - 1]
        url = row[COL_URL - 1]
        heb_dom = row[COL_HEB_DOM - 1]
        heb_site = row[COL_HEB_SITE - 1]
        resp = row[COL_RESP - 1]
        date_exp = row[COL_DATE_EXP - 1]

        if not url or not str(url).strip().startswith("http"):
            continue

        url = str(url).strip()
        domaine_str = str(domaine or "").strip()
        client = str(client or domaine_str or "Inconnu").strip()
        log.info(f"-> {client} | {url}")

        try:
            is_up, up_msg, response_ms, http_status = check_uptime(url, timeout)
            ssl_st, ssl_msg, ssl_days, ssl_issuer = check_ssl(url, ssl_warn)
            ndd_st, ndd_msg, ndd_days = check_domain(url, ndd_warn)
            stack_info = check_stack(url)
            psi = {}
            if run_psi:
                psi = check_pagespeed(url)
        except Exception as e:
            log.error(f"  Erreur inattendue sur {url} : {e} — site ignore, on continue")
            continue

        # Backup : lookup par domaine
        backup = backup_index.get(domaine_str.lower(), {})

        # Construction des alertes
        issues = []
        if not is_up:
            issues.append(("Site DOWN", up_msg, 1))
        if ssl_st == "critical":
            issues.append(("SSL CRITIQUE", ssl_msg, 1))
        elif ssl_st == "warning":
            issues.append(("SSL EXPIRATION", ssl_msg, 2))
        if ndd_st == "critical":
            issues.append(("NDD CRITIQUE", ndd_msg, 1))
        elif ndd_st == "warning":
            issues.append(("NDD EXPIRATION", ndd_msg, 2))
        if backup.get("status") == "critical":
            issues.append(("BACKUP MANQUANT", f"Dernier backup il y a {backup.get('days_since')}j", 2))
        # NOTE : PHP obsolete — pas d'email (éviterait de spammer chaque jour).
        # L'info reste visible dans le dashboard via stack_info.php_outdated.

        ticket_ids = []
        for title_prefix, detail, _ in issues:
            subject = f"[MONITORING] {title_prefix} - {client} ({url})"
            body = (
                f"Client    : {client}\n"
                f"URL       : {url}\n"
                f"Domaine   : {domaine_str or '?'}\n"
                f"Probleme  : {title_prefix}\n"
                f"Detail    : {detail}\n"
                f"Hebergeur : {heb_dom or '?'} / {heb_site or '?'}\n"
                f"Resp.     : {resp or '?'}\n"
                f"Detecte   : {now}\n"
            )
            ok = send_support_email(cfg, subject, body)
            if ok:
                ticket_ids.append("email envoye")
            log.warning(f"  {title_prefix}: {detail}")

        if not issues:
            log.info("  OK")

        r = {
            "client": client,
            "domaine": domaine_str,
            "url": url,
            "heb_dom": str(heb_dom or "-"),
            "heb_site": str(heb_site or "-"),
            "resp": str(resp or "-"),
            "date_exp": str(date_exp or "-"),
            "up": is_up,
            "up_msg": up_msg,
            "http_status": http_status,
            "response_ms": response_ms,
            "ssl_st": ssl_st,
            "ssl_msg": ssl_msg,
            "ssl_days": ssl_days,
            "ssl_issuer": ssl_issuer,
            "ndd_st": ndd_st,
            "ndd_msg": ndd_msg,
            "ndd_days": ndd_days,
            "psi": psi,
            "stack": stack_info,
            "backup": backup,
            "tickets": ticket_ids,
            "checked": now,
        }
        results.append(r)
        excel_writes.append((excel_row_num, r))

    wb.close()

    # Ecriture dans l Excel (K-P)
    _write_excel_results(excel_writes)

    # Historique SQLite
    n = history.save_run(HISTORY_DB, results)
    log.info(f"History : {n} lignes ajoutees dans {HISTORY_DB.name}")

    # Sauvegarde JSON pour l'interface web
    payload = {
        "sites": results,
        "last_run": now,
        "running": False,
        "psi_enabled": run_psi,
    }
    RESULTS_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    alertes_count = sum(
        1
        for r in results
        if not r["up"]
        or r["ssl_st"] in ("warning", "critical")
        or r["ndd_st"] in ("warning", "critical")
        or (r.get("backup") or {}).get("status") == "critical"
    )
    if alertes_count == 0:
        log.info(f"=== Termine - {len(results)} sites verifies — TOUT EST OK ===\n")
    else:
        log.info(f"=== Termine - {len(results)} sites verifies — {alertes_count} ALERTE(S) ===\n")

    if test_mode:
        log.info("Mode test : envoi du rapport complet...")
        send_test_report(results, cfg, now)


def _write_excel_results(excel_writes: list):
    """Ecrit les resultats dans les colonnes K-P de l'Excel."""
    if not excel_writes:
        return
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Inventaire"]
        for row_num, r in excel_writes:
            ws.cell(row=row_num, column=11).value = "UP" if r["up"] else "DOWN"
            ws.cell(row=row_num, column=12).value = r["ssl_msg"]
            ws.cell(row=row_num, column=13).value = (
                f"{r['ssl_days']}j" if r["ssl_days"] is not None else r["ssl_msg"]
            )
            ws.cell(row=row_num, column=14).value = r["ndd_msg"]
            ws.cell(row=row_num, column=15).value = r["checked"]
            ws.cell(row=row_num, column=16).value = (
                ", ".join(r["tickets"]) if r["tickets"] else ""
            )
        wb.save(EXCEL_PATH)
        wb.close()
        log.info(f"Excel mis a jour ({len(excel_writes)} lignes)")
    except PermissionError:
        log.error("Impossible d'ecrire dans l'Excel — fichier ouvert dans Excel ou OneDrive.")
    except Exception as e:
        log.error(f"Erreur mise a jour Excel : {e}")


if __name__ == "__main__":
    test_mode = "--test" in sys.argv
    run_psi = "--psi" in sys.argv or bool(os.environ.get("PSI_API_KEY"))
    if test_mode:
        log.info("=== MODE TEST — rapport envoye a axel.dos-santos@albys.com ===")
    if run_psi:
        log.info("=== PageSpeed Insights active ===")
    run(test_mode=test_mode, run_psi=run_psi)
